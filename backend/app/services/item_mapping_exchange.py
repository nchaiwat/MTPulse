from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import date
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import distinct, select
from sqlalchemy.orm import Session

from app.models import AuditEvent, BranchMapping, ItemMapping, ModernTrade, SalesInventoryFact

SHEET_NAME = "Item Mapping"
BRANCH_SHEET_NAME = "Branch Mapping"
HEADERS = (
    "TWD SKU",
    "TWD Description",
    "WA Item",
    "WA Description",
    "Mapping Status",
    "Import Note",
)
BRANCH_HEADERS = (
    "TWD Branch",
    "TWD Branch Description",
    "WA Branch",
    "WA Branch Description",
    "Mapping Status",
    "Import Note",
)


@dataclass(frozen=True)
class ExportItem:
    source_sku: str
    source_description: str
    wa_item_code: str
    wa_item_description: str
    status: str


@dataclass(frozen=True)
class ExportBranch:
    source_branch_code: str
    source_branch_description: str
    wa_branch_code: str
    wa_branch_description: str
    status: str


@dataclass(frozen=True)
class ImportCandidate:
    source_sku: str
    source_description: str
    wa_item_code: str
    wa_item_description: str


@dataclass(frozen=True)
class BranchImportCandidate:
    source_branch_code: str
    source_branch_description: str
    wa_branch_code: str
    wa_branch_description: str


@dataclass(frozen=True)
class ParsedItemWorkbook:
    row_count: int
    candidates: tuple[ImportCandidate, ...]
    skipped_blank: int
    conflicts: tuple[str, ...]
    errors: tuple[str, ...]
    branch_row_count: int
    branch_candidates: tuple[BranchImportCandidate, ...]
    branch_skipped_blank: int
    branch_conflicts: tuple[str, ...]


@dataclass(frozen=True)
class ItemImportReport:
    total_rows: int
    candidates: int
    inserted_pending: int
    unchanged: int
    skipped_blank: int
    new_source_skus: int
    conflicts: int
    branch_candidates: int
    branch_inserted_pending: int
    branch_updated: int
    branch_unchanged: int
    branch_skipped_blank: int
    branch_conflicts: int
    errors: tuple[str, ...]


def _text(value: Any) -> str:
    return "" if value is None else str(value).replace("\u00a0", " ").strip()


def _identifier(value: Any, width: int) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        raw = str(int(value))
    else:
        raw = _text(value)
        if raw.endswith(".0") and raw[:-2].isdigit():
            raw = raw[:-2]
    return raw.zfill(width) if raw.isdigit() else raw


def _set_text(cell, value: str) -> None:
    cell.value = value
    cell.data_type = "s"
    cell.number_format = "@"


def _style_mapping_sheet(sheet, row_count: int, table_name: str, widths: tuple[int, ...]) -> None:
    header_fill = PatternFill("solid", fgColor="0B756E")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    last_column = openpyxl.utils.get_column_letter(len(widths))
    sheet.auto_filter.ref = f"A1:{last_column}{max(1, row_count + 1)}"
    sheet.row_dimensions[1].height = 24
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    if row_count:
        table = Table(displayName=table_name, ref=f"A1:{last_column}{row_count + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def build_item_mapping_workbook(
    items: list[ExportItem], branches: list[ExportBranch] | None = None
) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)

    for row_number, item in enumerate(items, start=2):
        values = (
            item.source_sku,
            item.source_description,
            item.wa_item_code,
            item.wa_item_description,
            item.status,
            "กรอก WA Item และ WA Description แล้ว Import กลับเข้าระบบ"
            if not item.wa_item_code
            else "มี Mapping ในระบบแล้ว",
        )
        for column, value in enumerate(values, start=1):
            _set_text(sheet.cell(row=row_number, column=column), value)

    _style_mapping_sheet(sheet, len(items), "TWDItemMapping", (16, 48, 22, 48, 18, 52))

    branch_sheet = workbook.create_sheet(BRANCH_SHEET_NAME)
    branch_sheet.append(BRANCH_HEADERS)
    for row_number, branch in enumerate(branches or [], start=2):
        values = (
            branch.source_branch_code,
            branch.source_branch_description,
            branch.wa_branch_code,
            branch.wa_branch_description,
            branch.status,
            "กรอก WA Branch และ WA Branch Description แล้ว Import กลับเข้าระบบ"
            if not branch.wa_branch_code
            else "มี Mapping ในระบบแล้ว",
        )
        for column, value in enumerate(values, start=1):
            _set_text(branch_sheet.cell(row=row_number, column=column), value)
    _style_mapping_sheet(
        branch_sheet, len(branches or []), "TWDBranchMapping", (16, 52, 22, 42, 18, 54)
    )

    instructions = workbook.create_sheet("วิธีใช้งาน")
    instructions.column_dimensions["A"].width = 110
    notes = (
        "1. ใช้ Sheet 'Item Mapping' สำหรับ VLOOKUP และแก้ไขข้อมูล",
        "2. ห้ามแก้ TWD SKU เพราะเป็น Key ที่ใช้ Import",
        "3. กรอก WA Item และ WA Description เฉพาะรายการที่ยังไม่มี Mapping",
        "4. สามารถเพิ่ม TWD SKU และ TWD Description ใหม่ต่อท้ายตารางได้",
        "5. Import จะไม่แก้ทับ Mapping เดิม และรายการใหม่จะเข้าระบบเป็นสถานะรอตรวจสอบ",
        "6. หากใช้สูตร VLOOKUP ให้เปิดไฟล์และ Save ด้วย Excel ก่อน Import เพื่อบันทึกค่าที่คำนวณแล้ว",
        "7. ใช้ Sheet 'Branch Mapping' เพื่อกรอก WA Branch และชื่อ Branch ที่ต้องการแสดงในระบบ",
    )
    for row, note in enumerate(notes, start=1):
        _set_text(instructions.cell(row=row, column=1), note)
    instructions["A1"].font = Font(bold=True, color="0B756E")

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def parse_item_mapping_workbook(content: bytes) -> ParsedItemWorkbook:
    try:
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("ไฟล์ Excel ไม่ถูกต้องหรือเปิดอ่านไม่ได้") from exc

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"ไม่พบ Sheet '{SHEET_NAME}'")
        sheet = workbook[SHEET_NAME]
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header_map = {_text(value): index for index, value in enumerate(first_row)}
        required = ("TWD SKU", "WA Item")
        missing = [header for header in required if header not in header_map]
        if missing:
            raise ValueError(f"ไม่พบ Column ที่จำเป็น: {', '.join(missing)}")

        candidate_sets: dict[str, set[tuple[str, str, str]]] = {}
        errors: list[str] = []
        row_count = skipped_blank = 0
        description_index = header_map.get("WA Description")
        source_description_index = header_map.get("TWD Description")
        for excel_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            row_count += 1
            sku = _identifier(row[header_map["TWD SKU"]], 9)
            wa_item = _text(row[header_map["WA Item"]])
            source_description = (
                _text(row[source_description_index])
                if source_description_index is not None and source_description_index < len(row)
                else ""
            )
            wa_description = (
                _text(row[description_index])
                if description_index is not None and description_index < len(row)
                else ""
            )
            if not sku:
                errors.append(f"แถว {excel_row}: ไม่มี TWD SKU")
                continue
            if not wa_item:
                skipped_blank += 1
                continue
            if len(sku) > 50 or len(wa_item) > 50:
                errors.append(f"แถว {excel_row}: รหัสยาวเกิน 50 ตัวอักษร")
                continue
            candidate_sets.setdefault(sku, set()).add(
                (source_description, wa_item, wa_description)
            )

        conflicts = tuple(sorted(sku for sku, values in candidate_sets.items() if len(values) > 1))
        candidates = tuple(
            ImportCandidate(sku, *next(iter(values)))
            for sku, values in sorted(candidate_sets.items())
            if len(values) == 1
        )

        branch_sets: dict[str, set[tuple[str, str, str]]] = {}
        branch_row_count = branch_skipped_blank = 0
        if BRANCH_SHEET_NAME in workbook.sheetnames:
            branch_sheet = workbook[BRANCH_SHEET_NAME]
            branch_first_row = next(
                branch_sheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
            )
            branch_headers = {
                _text(value): index for index, value in enumerate(branch_first_row)
            }
            branch_required = ("TWD Branch", "WA Branch")
            branch_missing = [
                header for header in branch_required if header not in branch_headers
            ]
            if branch_missing:
                raise ValueError(
                    f"Sheet '{BRANCH_SHEET_NAME}' ไม่พบ Column: {', '.join(branch_missing)}"
                )
            source_name_index = branch_headers.get("TWD Branch Description")
            wa_name_index = branch_headers.get("WA Branch Description")
            for excel_row, row in enumerate(
                branch_sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if not any(value not in (None, "") for value in row):
                    continue
                branch_row_count += 1
                source_code = _identifier(row[branch_headers["TWD Branch"]], 5)
                wa_code = _text(row[branch_headers["WA Branch"]])
                source_name = (
                    _text(row[source_name_index])
                    if source_name_index is not None and source_name_index < len(row)
                    else ""
                )
                wa_name = (
                    _text(row[wa_name_index])
                    if wa_name_index is not None and wa_name_index < len(row)
                    else ""
                )
                if not source_code:
                    errors.append(f"Branch แถว {excel_row}: ไม่มี TWD Branch")
                    continue
                if not wa_code:
                    branch_skipped_blank += 1
                    continue
                if len(source_code) > 30 or len(wa_code) > 30:
                    errors.append(f"Branch แถว {excel_row}: รหัสยาวเกิน 30 ตัวอักษร")
                    continue
                branch_sets.setdefault(source_code, set()).add(
                    (source_name, wa_code, wa_name)
                )

        branch_conflicts = tuple(
            sorted(code for code, values in branch_sets.items() if len(values) > 1)
        )
        branch_candidates = tuple(
            BranchImportCandidate(code, *next(iter(values)))
            for code, values in sorted(branch_sets.items())
            if len(values) == 1
        )
        return ParsedItemWorkbook(
            row_count=row_count,
            candidates=candidates,
            skipped_blank=skipped_blank,
            conflicts=conflicts,
            errors=tuple(errors[:50]),
            branch_row_count=branch_row_count,
            branch_candidates=branch_candidates,
            branch_skipped_blank=branch_skipped_blank,
            branch_conflicts=branch_conflicts,
        )
    finally:
        workbook.close()


def import_item_mapping_workbook(
    session: Session,
    content: bytes,
    effective_from: date,
    filename: str,
) -> ItemImportReport:
    parsed = parse_item_mapping_workbook(content)
    modern_trade = session.scalar(select(ModernTrade).where(ModernTrade.code == "TWD"))
    if modern_trade is None:
        raise ValueError("ยังไม่มี Modern Trade รหัส TWD ในฐานข้อมูล")

    source_skus = set(session.scalars(select(distinct(SalesInventoryFact.source_sku))).all())
    source_branches = set(
        session.scalars(select(distinct(SalesInventoryFact.source_branch_code))).all()
    )
    active_mappings = session.scalars(
        select(ItemMapping)
        .where(
            ItemMapping.modern_trade_id == modern_trade.id,
            ItemMapping.effective_from <= effective_from,
            (ItemMapping.effective_to.is_(None) | (ItemMapping.effective_to >= effective_from)),
        )
        .order_by(ItemMapping.effective_from)
    ).all()
    existing = {mapping.source_sku: mapping for mapping in active_mappings}
    active_branch_mappings = session.scalars(
        select(BranchMapping)
        .where(
            BranchMapping.modern_trade_id == modern_trade.id,
            BranchMapping.effective_from <= effective_from,
            (
                BranchMapping.effective_to.is_(None)
                | (BranchMapping.effective_to >= effective_from)
            ),
        )
        .order_by(BranchMapping.effective_from)
    ).all()
    existing_branches = {
        mapping.source_branch_code: mapping for mapping in active_branch_mappings
    }
    actor = f"excel-import:{filename[:160]}"
    inserted = unchanged = new_source_skus = existing_conflicts = 0
    errors = list(parsed.errors)

    for candidate in parsed.candidates:
        current = existing.get(candidate.source_sku)
        if current:
            if current.wa_item_code == candidate.wa_item_code:
                unchanged += 1
            else:
                existing_conflicts += 1
                errors.append(
                    f"TWD SKU {candidate.source_sku} มี Mapping เดิม {current.wa_item_code}; ไม่ได้แก้ทับ"
                )
            continue

        if candidate.source_sku not in source_skus:
            new_source_skus += 1

        mapping = ItemMapping(
            modern_trade_id=modern_trade.id,
            source_sku=candidate.source_sku,
            source_description=candidate.source_description or None,
            wa_item_code=candidate.wa_item_code,
            wa_item_description=candidate.wa_item_description or None,
            status="pending",
            effective_from=effective_from,
            effective_to=None,
            changed_by=actor,
        )
        session.add(mapping)
        session.add(
            AuditEvent(
                entity_type="item_mapping",
                entity_id=candidate.source_sku,
                action="import_pending_mapping",
                actor=actor,
                before_json=None,
                after_json=json.dumps(
                    {**asdict(candidate), "effective_from": effective_from.isoformat()},
                    ensure_ascii=False,
                ),
            )
        )
        inserted += 1

    branch_inserted = branch_updated = branch_unchanged = branch_existing_conflicts = 0
    for candidate in parsed.branch_candidates:
        current = existing_branches.get(candidate.source_branch_code)
        if current:
            if current.wa_branch_code == candidate.wa_branch_code:
                next_source_description = (
                    candidate.source_branch_description
                    or current.source_branch_description
                )
                next_wa_description = (
                    candidate.wa_branch_description or current.wa_branch_description
                )
                if (
                    next_source_description != current.source_branch_description
                    or next_wa_description != current.wa_branch_description
                ):
                    before = {
                        "source_branch_description": current.source_branch_description,
                        "wa_branch_description": current.wa_branch_description,
                    }
                    current.source_branch_description = next_source_description
                    current.wa_branch_description = next_wa_description
                    current.changed_by = actor
                    session.add(
                        AuditEvent(
                            entity_type="branch_mapping",
                            entity_id=candidate.source_branch_code,
                            action="update_mapping_descriptions",
                            actor=actor,
                            before_json=json.dumps(before, ensure_ascii=False),
                            after_json=json.dumps(asdict(candidate), ensure_ascii=False),
                        )
                    )
                    branch_updated += 1
                else:
                    branch_unchanged += 1
            else:
                branch_existing_conflicts += 1
                errors.append(
                    f"TWD Branch {candidate.source_branch_code} มี Mapping เดิม "
                    f"{current.wa_branch_code}; ไม่ได้แก้ทับ"
                )
            continue

        mapping = BranchMapping(
            modern_trade_id=modern_trade.id,
            source_branch_code=candidate.source_branch_code,
            source_branch_description=candidate.source_branch_description or None,
            wa_branch_code=candidate.wa_branch_code,
            wa_branch_description=candidate.wa_branch_description or None,
            status="pending",
            effective_from=effective_from,
            effective_to=None,
            changed_by=actor,
        )
        session.add(mapping)
        session.add(
            AuditEvent(
                entity_type="branch_mapping",
                entity_id=candidate.source_branch_code,
                action="import_pending_mapping",
                actor=actor,
                before_json=None,
                after_json=json.dumps(
                    {
                        **asdict(candidate),
                        "effective_from": effective_from.isoformat(),
                        "source_seen": candidate.source_branch_code in source_branches,
                    },
                    ensure_ascii=False,
                ),
            )
        )
        branch_inserted += 1

    session.commit()
    return ItemImportReport(
        total_rows=parsed.row_count,
        candidates=len(parsed.candidates),
        inserted_pending=inserted,
        unchanged=unchanged,
        skipped_blank=parsed.skipped_blank,
        new_source_skus=new_source_skus,
        conflicts=len(parsed.conflicts) + existing_conflicts,
        branch_candidates=len(parsed.branch_candidates),
        branch_inserted_pending=branch_inserted,
        branch_updated=branch_updated,
        branch_unchanged=branch_unchanged,
        branch_skipped_blank=parsed.branch_skipped_blank,
        branch_conflicts=len(parsed.branch_conflicts) + branch_existing_conflicts,
        errors=tuple(errors[:50]),
    )
