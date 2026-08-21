from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import distinct, select
from sqlalchemy.orm import Session

from app.models import (
    AuditEvent,
    BranchMapping,
    ItemMapping,
    ModernTrade,
    SalesInventoryFact,
)


@dataclass(frozen=True)
class ItemCandidate:
    source_sku: str
    wa_item_code: str
    wa_item_description: str


@dataclass(frozen=True)
class BranchCandidate:
    source_branch_code: str
    wa_branch_code: str


@dataclass(frozen=True)
class MappingExtract:
    items: tuple[ItemCandidate, ...]
    branches: tuple[BranchCandidate, ...]
    raw_item_rows: int
    duplicate_item_rows: int
    item_conflicts: tuple[str, ...]
    unresolved_branch_rows: int


@dataclass(frozen=True)
class MappingImportReport:
    workbook_items: int
    workbook_branches: int
    item_conflicts: int
    unresolved_branch_rows: int
    source_items: int
    source_branches: int
    eligible_items: int
    eligible_branches: int
    inserted_items: int
    inserted_branches: int
    unchanged_items: int
    unchanged_branches: int
    existing_item_conflicts: int
    existing_branch_conflicts: int


def _text(value: Any) -> str:
    return "" if value is None else str(value).replace("\u00a0", " ").strip()


def _identifier(value: Any, width: int) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        raw = str(int(value))
    else:
        raw = _text(value)
        if raw.endswith(".0") and raw[:-2].isdigit():
            raw = raw[:-2]
    return raw.zfill(width) if raw.isdigit() else raw


def extract_mapping_workbook(path: str | Path) -> MappingExtract:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        item_sets: dict[str, set[tuple[str, str]]] = {}
        raw_item_rows = 0
        for row in workbook["BP 22.1.26"].iter_rows(min_row=2, values_only=True):
            if not _text(row[0]).startswith("CTW-"):
                continue
            raw_item_rows += 1
            source_sku = _identifier(row[4], 9)
            wa_item_code = _text(row[2])
            if source_sku and wa_item_code:
                item_sets.setdefault(source_sku, set()).add((wa_item_code, _text(row[3])))

        item_conflicts = tuple(sorted(sku for sku, values in item_sets.items() if len(values) > 1))
        items = tuple(
            ItemCandidate(sku, *next(iter(values)))
            for sku, values in sorted(item_sets.items())
            if len(values) == 1
        )

        branches: list[BranchCandidate] = []
        unresolved_branch_rows = 0
        for row in workbook["สาขา"].iter_rows(min_row=2, values_only=True):
            source_code = _identifier(row[1], 5)
            if not source_code:
                continue
            wa_branch_code = _text(row[3])
            if not wa_branch_code:
                unresolved_branch_rows += 1
                continue
            branches.append(BranchCandidate(source_code, wa_branch_code))
    finally:
        workbook.close()

    unique_branches = {candidate.source_branch_code: candidate for candidate in branches}
    return MappingExtract(
        items=items,
        branches=tuple(unique_branches[key] for key in sorted(unique_branches)),
        raw_item_rows=raw_item_rows,
        duplicate_item_rows=raw_item_rows - sum(len(values) for values in item_sets.values()),
        item_conflicts=item_conflicts,
        unresolved_branch_rows=unresolved_branch_rows,
    )


def import_mapping_workbook(
    session: Session,
    path: str | Path,
    effective_from: date,
    *,
    apply: bool,
) -> MappingImportReport:
    extract = extract_mapping_workbook(path)
    modern_trade = session.scalar(select(ModernTrade).where(ModernTrade.code == "TWD"))
    if modern_trade is None:
        raise ValueError("ยังไม่มี Modern Trade รหัส TWD ในฐานข้อมูล")

    source_items = set(session.scalars(select(distinct(SalesInventoryFact.source_sku))).all())
    source_branches = set(
        session.scalars(select(distinct(SalesInventoryFact.source_branch_code))).all()
    )
    eligible_items = [item for item in extract.items if item.source_sku in source_items]
    eligible_branches = [
        branch for branch in extract.branches if branch.source_branch_code in source_branches
    ]

    active_item_mappings = session.scalars(
        select(ItemMapping).where(
            ItemMapping.modern_trade_id == modern_trade.id,
            ItemMapping.effective_from <= effective_from,
            (ItemMapping.effective_to.is_(None) | (ItemMapping.effective_to >= effective_from)),
        )
    ).all()
    active_branch_mappings = session.scalars(
        select(BranchMapping).where(
            BranchMapping.modern_trade_id == modern_trade.id,
            BranchMapping.effective_from <= effective_from,
            (BranchMapping.effective_to.is_(None) | (BranchMapping.effective_to >= effective_from)),
        )
    ).all()
    existing_items = {mapping.source_sku: mapping for mapping in active_item_mappings}
    existing_branches = {
        mapping.source_branch_code: mapping for mapping in active_branch_mappings
    }

    inserted_items = unchanged_items = existing_item_conflicts = 0
    for candidate in eligible_items:
        existing = existing_items.get(candidate.source_sku)
        if existing:
            if existing.wa_item_code == candidate.wa_item_code:
                unchanged_items += 1
            else:
                existing_item_conflicts += 1
            continue
        inserted_items += 1
        if apply:
            mapping = ItemMapping(
                modern_trade_id=modern_trade.id,
                source_sku=candidate.source_sku,
                wa_item_code=candidate.wa_item_code,
                wa_item_description=candidate.wa_item_description or None,
                status="confirmed",
                effective_from=effective_from,
                effective_to=None,
                changed_by="workbook:KPI - TWD 2026.xlsx",
            )
            session.add(mapping)
            session.add(
                _audit("item_mapping", candidate.source_sku, asdict(candidate), effective_from)
            )

    inserted_branches = unchanged_branches = existing_branch_conflicts = 0
    for candidate in eligible_branches:
        existing = existing_branches.get(candidate.source_branch_code)
        if existing:
            if existing.wa_branch_code == candidate.wa_branch_code:
                unchanged_branches += 1
            else:
                existing_branch_conflicts += 1
            continue
        inserted_branches += 1
        if apply:
            mapping = BranchMapping(
                modern_trade_id=modern_trade.id,
                source_branch_code=candidate.source_branch_code,
                wa_branch_code=candidate.wa_branch_code,
                status="confirmed",
                effective_from=effective_from,
                effective_to=None,
                changed_by="workbook:KPI - TWD 2026.xlsx",
            )
            session.add(mapping)
            session.add(
                _audit(
                    "branch_mapping",
                    candidate.source_branch_code,
                    asdict(candidate),
                    effective_from,
                )
            )

    if apply:
        session.commit()

    return MappingImportReport(
        workbook_items=len(extract.items),
        workbook_branches=len(extract.branches),
        item_conflicts=len(extract.item_conflicts),
        unresolved_branch_rows=extract.unresolved_branch_rows,
        source_items=len(source_items),
        source_branches=len(source_branches),
        eligible_items=len(eligible_items),
        eligible_branches=len(eligible_branches),
        inserted_items=inserted_items,
        inserted_branches=inserted_branches,
        unchanged_items=unchanged_items,
        unchanged_branches=unchanged_branches,
        existing_item_conflicts=existing_item_conflicts,
        existing_branch_conflicts=existing_branch_conflicts,
    )


def _audit(entity_type: str, entity_id: str, value: dict[str, Any], effective_from: date):
    return AuditEvent(
        entity_type=entity_type,
        entity_id=entity_id,
        action="import_confirmed_mapping",
        actor="workbook:KPI - TWD 2026.xlsx",
        before_json=None,
        after_json=json.dumps(
            {**value, "effective_from": effective_from.isoformat()}, ensure_ascii=False
        ),
    )
