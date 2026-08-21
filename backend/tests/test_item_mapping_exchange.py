from datetime import date
from io import BytesIO

import openpyxl

from app.models import AuditEvent, BranchMapping, ItemMapping, ModernTrade
from app.services.item_mapping_exchange import (
    ExportBranch,
    ExportItem,
    build_item_mapping_workbook,
    import_item_mapping_workbook,
    parse_item_mapping_workbook,
)


def test_export_preserves_sku_as_text_and_includes_all_columns() -> None:
    content = build_item_mapping_workbook(
        [ExportItem("060424005", "สินค้าทดสอบ", "", "", "unmatched")]
    )

    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet = workbook["Item Mapping"]
    assert [cell.value for cell in sheet[1]] == [
        "TWD SKU",
        "TWD Description",
        "WA Item",
        "WA Description",
        "Mapping Status",
        "Import Note",
    ]
    assert sheet["A2"].value == "060424005"
    assert sheet["A2"].data_type == "s"
    assert sheet.freeze_panes == "A2"
    workbook.close()


def test_export_and_parse_branch_mapping_sheet() -> None:
    content = build_item_mapping_workbook(
        [],
        [ExportBranch("60001", "CRC Head Office", "WA-BKK", "สำนักงานใหญ่", "pending")],
    )

    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet = workbook["Branch Mapping"]
    assert sheet["A2"].value == "60001"
    assert sheet["A2"].data_type == "s"
    assert sheet["D2"].value == "สำนักงานใหญ่"
    workbook.close()

    parsed = parse_item_mapping_workbook(content)
    assert len(parsed.branch_candidates) == 1
    assert parsed.branch_candidates[0].source_branch_code == "60001"
    assert parsed.branch_candidates[0].wa_branch_code == "WA-BKK"


def test_parse_deduplicates_rows_and_rejects_conflicting_mapping() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Item Mapping"
    sheet.append(["TWD SKU", "TWD Description", "WA Item", "WA Description"])
    sheet.append([60424005, "สินค้า 1", "WA-001", "รายละเอียด 1"])
    sheet.append([60424005, "สินค้า 1", "WA-001", "รายละเอียด 1"])
    sheet.append([60424006, "สินค้า 2", "WA-002", "รายละเอียด 2"])
    sheet.append([60424006, "สินค้า 2", "WA-003", "รายละเอียด 3"])
    sheet.append([60424007, "สินค้า 3", None, None])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    parsed = parse_item_mapping_workbook(output.getvalue())

    assert parsed.row_count == 5
    assert parsed.skipped_blank == 1
    assert parsed.conflicts == ("060424006",)
    assert [(item.source_sku, item.wa_item_code) for item in parsed.candidates] == [
        ("060424005", "WA-001")
    ]


class _ScalarRows:
    def __init__(self, values: list[object]) -> None:
        self.values = values

    def all(self) -> list[object]:
        return self.values


class _ImportSession:
    def __init__(self) -> None:
        self.scalar_calls = 0
        self.scalars_calls = 0
        self.added: list[object] = []
        self.committed = False

    def scalar(self, _statement):
        self.scalar_calls += 1
        return ModernTrade(id=1, code="TWD", name="Thai Watsadu")

    def scalars(self, _statement) -> _ScalarRows:
        self.scalars_calls += 1
        return _ScalarRows([])

    def add(self, value: object) -> None:
        self.added.append(value)

    def commit(self) -> None:
        self.committed = True


class _ExistingBranchSession(_ImportSession):
    def __init__(self, branch_mapping: BranchMapping) -> None:
        super().__init__()
        self.branch_mapping = branch_mapping

    def scalars(self, _statement) -> _ScalarRows:
        self.scalars_calls += 1
        values_by_call = {
            1: [],
            2: [self.branch_mapping.source_branch_code],
            3: [],
            4: [self.branch_mapping],
        }
        return _ScalarRows(values_by_call[self.scalars_calls])


def test_import_accepts_new_source_sku_as_pending_mapping() -> None:
    content = build_item_mapping_workbook(
        [ExportItem("099999999", "สินค้าใหม่", "WA-NEW", "รายละเอียดใหม่", "unmatched")]
    )
    session = _ImportSession()

    report = import_item_mapping_workbook(
        session, content, date(2026, 8, 16), "new-item.xlsx"  # type: ignore[arg-type]
    )

    mapping = next(value for value in session.added if isinstance(value, ItemMapping))
    assert mapping.source_sku == "099999999"
    assert mapping.source_description == "สินค้าใหม่"
    assert mapping.status == "pending"
    assert any(isinstance(value, AuditEvent) for value in session.added)
    assert report.inserted_pending == 1
    assert report.new_source_skus == 1
    assert report.errors == ()
    assert session.committed


def test_import_accepts_branch_mapping_as_pending() -> None:
    content = build_item_mapping_workbook(
        [],
        [ExportBranch("69999", "สาขาต้นทาง", "WA-NEW", "สาขาใหม่", "unmatched")],
    )
    session = _ImportSession()

    report = import_item_mapping_workbook(
        session, content, date(2026, 8, 16), "new-branch.xlsx"  # type: ignore[arg-type]
    )

    mapping = next(value for value in session.added if isinstance(value, BranchMapping))
    assert mapping.source_branch_code == "69999"
    assert mapping.wa_branch_description == "สาขาใหม่"
    assert mapping.status == "pending"
    assert report.branch_inserted_pending == 1
    assert report.branch_conflicts == 0


def test_import_updates_branch_description_without_changing_existing_code() -> None:
    existing = BranchMapping(
        modern_trade_id=1,
        source_branch_code="60001",
        source_branch_description="CRC Head Office",
        wa_branch_code="WA-BKK",
        wa_branch_description=None,
        status="confirmed",
        effective_from=date(2026, 8, 16),
        effective_to=None,
        changed_by="original",
    )
    content = build_item_mapping_workbook(
        [],
        [ExportBranch("60001", "CRC Head Office", "WA-BKK", "สำนักงานใหญ่", "confirmed")],
    )
    session = _ExistingBranchSession(existing)

    report = import_item_mapping_workbook(
        session, content, date(2026, 8, 16), "branch-name.xlsx"  # type: ignore[arg-type]
    )

    assert existing.wa_branch_code == "WA-BKK"
    assert existing.wa_branch_description == "สำนักงานใหญ่"
    assert report.branch_updated == 1
    assert report.branch_unchanged == 0
